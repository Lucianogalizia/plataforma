# ==========================================================
# backend/api/rrhh.py
#
# Endpoints REST para el módulo RRHH / Guardias
# Cache en memoria (TTLCache) para respuesta instantánea.
#
# Estrategia:
#   GETs  → cachean resultado con TTL
#   POSTs → invalidan claves afectadas y retornan dato fresco
# ==========================================================

from __future__ import annotations

import io
from typing import List, Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from core import rrhh_db as db
from core.cache import cache

router = APIRouter()

# ==========================================================
# TTLs
# ==========================================================
_TTL_PERSONAL  = 1800   # 30 min  — cambia poco
_TTL_PERIODOS  = 3600   # 1 hora  — estático dentro del día
_TTL_PARTE     = 300    # 5 min   — cambia al guardar/enviar/aprobar
_TTL_BITACORA  = 300    # 5 min
_TTL_PENDIENTE = 120    # 2 min   — el líder necesita ver novedades rápido
_TTL_CONSOL    = 300    # 5 min


# ==========================================================
# Helpers de caché
# ==========================================================

def _k_parte(legajo: str, periodo: str)    -> str: return f"rrhh:parte:{legajo}:{periodo}"
def _k_bitacora(legajo: str)               -> str: return f"rrhh:bitacora:{legajo}"
def _k_pendientes(leader: str)             -> str: return f"rrhh:pendientes:{leader}"
def _k_consolidado(leader: str, per: str)  -> str: return f"rrhh:consolidado:{leader}:{per}"
_K_PERSONAL = "rrhh:personal"
_K_PERIODOS = "rrhh:periodos"


def _invalidar_parte(legajo: str, periodo: str, leader_legajo: Optional[str] = None):
    """Invalida todas las claves afectadas por un cambio en un parte."""
    cache.delete(_k_parte(legajo, periodo))
    cache.delete(_k_bitacora(legajo))
    # Buscar leader si no se pasó
    if not leader_legajo:
        p = db.get_person(legajo)
        leader_legajo = p.get("leader_legajo") if p else None
    if leader_legajo:
        cache.delete(_k_pendientes(leader_legajo))
        cache.delete(_k_consolidado(leader_legajo, periodo))


# ==========================================================
# Modelos Pydantic
# ==========================================================

class LoginBody(BaseModel):
    legajo: str
    cuil:   str

class ItemIn(BaseModel):
    fecha:      str
    tipo:       str
    valor_num:  Optional[float] = None
    comentario: Optional[str]  = None

class SaveParteBody(BaseModel):
    items: List[ItemIn]

class AprobarBody(BaseModel):
    aprobador_legajo: str

class RechazarBody(BaseModel):
    comentario:       str
    aprobador_legajo: str

class ImportPersonalRow(BaseModel):
    legajo:        str
    cuil:          str
    nombre:        str
    leader_legajo: str
    funcion:       Optional[str] = None
    origen:        Optional[str] = None
    lugar_trabajo: Optional[str] = None

class ImportPersonalBody(BaseModel):
    rows: List[ImportPersonalRow]


# ==========================================================
# Login  (sin cache — siempre fresco)
# ==========================================================

@router.post("/login")
async def login(body: LoginBody):
    ok, user, err = db.verify_login(body.legajo, body.cuil)
    if not ok:
        raise HTTPException(status_code=401, detail=err)
    return {"ok": True, "user": user}


# ==========================================================
# Períodos
# ==========================================================

@router.get("/periodos")
async def get_periodos(n: int = 8):
    cached = cache.get(_K_PERIODOS)
    if cached is not None:
        return cached
    result = {
        "actual":   db.current_period_id(),
        "periodos": db.recent_periods(n),
    }
    cache.set(_K_PERIODOS, result, ttl=_TTL_PERIODOS)
    return result


# ==========================================================
# Personal
# ==========================================================

@router.get("/personal")
async def get_personal():
    cached = cache.get(_K_PERSONAL)
    if cached is not None:
        return cached
    result = {"personal": db.list_personal()}
    cache.set(_K_PERSONAL, result, ttl=_TTL_PERSONAL)
    return result


@router.post("/personal/import")
async def import_personal(body: ImportPersonalBody):
    rows = [r.model_dump() for r in body.rows]
    ins, upd = db.upsert_personal(rows)
    cache.delete(_K_PERSONAL)           # invalidar lista
    return {"ok": True, "insertados": ins, "actualizados": upd}


# ==========================================================
# Parte — construcción de respuesta completa
# ==========================================================

def _build_parte_response(legajo: str, periodo: str) -> dict:
    parte  = db.get_or_create_parte(legajo, periodo)
    items  = db.list_items(legajo, periodo)
    dates  = db.period_dates(periodo)
    start, end = db.period_bounds(periodo)

    by_date: dict = {}
    for it in items:
        f, t = it["fecha"], it["tipo"]
        if f not in by_date:
            by_date[f] = {"G":False,"F":False,"D":False,"HO":False,
                          "HV":0.0,"HE":0.0,"comentario":""}
        if t in db.TIPOS_DIA:
            by_date[f][t] = True
        elif t == "HV":
            by_date[f]["HV"] = round(float(it.get("valor_num") or 0), 2)
        elif t == "HE":
            by_date[f]["HE"] = round(float(it.get("valor_num") or 0), 2)
        if it.get("comentario") and not by_date[f]["comentario"]:
            by_date[f]["comentario"] = it["comentario"]

    grilla = []
    totales = {t: 0 for t in db.TIPOS_DIA}
    totales["HV"] = totales["HE"] = 0.0

    for d in dates:
        f   = d.isoformat()
        row = by_date.get(f, {"G":False,"F":False,"D":False,"HO":False,
                               "HV":0.0,"HE":0.0,"comentario":""})
        grilla.append({"fecha": f, **row})
        for t in db.TIPOS_DIA:
            if row[t]: totales[t] += 1
        totales["HV"] = round(totales["HV"] + row["HV"], 2)
        totales["HE"] = round(totales["HE"] + row["HE"], 2)

    return {
        "legajo":            legajo,
        "periodo":           periodo,
        "periodo_display":   db.period_display(periodo),
        "periodo_inicio":    start.isoformat(),
        "periodo_fin":       end.isoformat(),
        "estado":            parte["estado"],
        "submitted_at":      parte.get("submitted_at"),
        "approved_at":       parte.get("approved_at"),
        "approved_by":       parte.get("approved_by_legajo"),
        "rejection_comment": parte.get("rejection_comment"),
        "grilla":            grilla,
        "totales":           totales,
    }


# ==========================================================
# Parte — Leer
# ==========================================================

@router.get("/parte/{legajo}/{periodo}")
async def get_parte(legajo: str, periodo: str):
    key    = _k_parte(legajo, periodo)
    cached = cache.get(key)
    if cached is not None:
        return cached
    result = _build_parte_response(legajo, periodo)
    cache.set(key, result, ttl=_TTL_PARTE)
    return result


# ==========================================================
# Parte — Guardar borrador
# ==========================================================

@router.post("/parte/{legajo}/{periodo}/guardar")
async def guardar_parte(legajo: str, periodo: str, body: SaveParteBody):
    parte = db.get_or_create_parte(legajo, periodo)
    if parte["estado"] not in ("BORRADOR", "RECHAZADO"):
        raise HTTPException(400, f"No se puede editar un parte en estado {parte['estado']}.")
    db.save_items(legajo, periodo, [it.model_dump() for it in body.items])
    db.update_parte_estado(legajo, periodo, "BORRADOR", rejection_comment=None)
    _invalidar_parte(legajo, periodo)
    result = _build_parte_response(legajo, periodo)
    cache.set(_k_parte(legajo, periodo), result, ttl=_TTL_PARTE)
    return {"ok": True, "estado": "BORRADOR", "parte": result}


# ==========================================================
# Parte — Enviar a aprobación
# ==========================================================

@router.post("/parte/{legajo}/{periodo}/enviar")
async def enviar_parte(legajo: str, periodo: str, body: SaveParteBody):
    parte = db.get_or_create_parte(legajo, periodo)
    if parte["estado"] not in ("BORRADOR", "RECHAZADO"):
        raise HTTPException(400, f"No se puede enviar un parte en estado {parte['estado']}.")
    db.save_items(legajo, periodo, [it.model_dump() for it in body.items])
    db.update_parte_estado(legajo, periodo, "ENVIADO",
                           submitted_at=db.utcnow_str(), rejection_comment=None)
    _invalidar_parte(legajo, periodo)
    result = _build_parte_response(legajo, periodo)
    cache.set(_k_parte(legajo, periodo), result, ttl=_TTL_PARTE)
    return {"ok": True, "estado": "ENVIADO", "parte": result}


# ==========================================================
# Parte líder — guardar propio (auto-aprobado)
# ==========================================================

@router.post("/parte/{legajo}/{periodo}/guardar-lider")
async def guardar_parte_lider(legajo: str, periodo: str, body: SaveParteBody):
    db.get_or_create_parte(legajo, periodo)
    db.save_items(legajo, periodo, [it.model_dump() for it in body.items])
    db.update_parte_estado(legajo, periodo, "APROBADO",
                           submitted_at=db.utcnow_str(),
                           approved_at=db.utcnow_str(),
                           approved_by_legajo=legajo,
                           rejection_comment=None)
    _invalidar_parte(legajo, periodo, leader_legajo=legajo)
    result = _build_parte_response(legajo, periodo)
    cache.set(_k_parte(legajo, periodo), result, ttl=_TTL_PARTE)
    return {"ok": True, "estado": "APROBADO", "parte": result}


# ==========================================================
# Aprobación (líder)
# ==========================================================

@router.post("/parte/{legajo}/{periodo}/aprobar")
async def aprobar_parte(legajo: str, periodo: str, body: AprobarBody):
    parte = db.get_parte(legajo, periodo)
    if not parte:
        raise HTTPException(404, "Parte no encontrado.")
    if parte["estado"] != "ENVIADO":
        raise HTTPException(400, f"Solo se pueden aprobar partes ENVIADOS. Estado: {parte['estado']}.")
    person = db.get_person(legajo)
    aprobador = str(body.aprobador_legajo).strip()
    if not person or (aprobador not in db.SUPER_LIDERES and str(person.get("leader_legajo","")).strip() != aprobador):
        raise HTTPException(403, "No tenés permiso para aprobar este parte.")
    db.update_parte_estado(legajo, periodo, "APROBADO",
                           approved_at=db.utcnow_str(),
                           approved_by_legajo=body.aprobador_legajo,
                           rejection_comment=None)
    _invalidar_parte(legajo, periodo, leader_legajo=body.aprobador_legajo)
    return {"ok": True, "estado": "APROBADO"}


# ==========================================================
# Rechazo (líder)
# ==========================================================

@router.post("/parte/{legajo}/{periodo}/rechazar")
async def rechazar_parte(legajo: str, periodo: str, body: RechazarBody):
    if not body.comentario.strip():
        raise HTTPException(400, "El comentario es obligatorio al rechazar.")
    parte = db.get_parte(legajo, periodo)
    if not parte:
        raise HTTPException(404, "Parte no encontrado.")
    if parte["estado"] != "ENVIADO":
        raise HTTPException(400, f"Solo se pueden rechazar partes ENVIADOS. Estado: {parte['estado']}.")
    person = db.get_person(legajo)
    aprobador = str(body.aprobador_legajo).strip()
    if not person or (aprobador not in db.SUPER_LIDERES and str(person.get("leader_legajo","")).strip() != aprobador):
        raise HTTPException(403, "No tenés permiso para rechazar este parte.")
    db.update_parte_estado(legajo, periodo, "RECHAZADO",
                           approved_by_legajo=body.aprobador_legajo,
                           rejection_comment=body.comentario.strip())
    _invalidar_parte(legajo, periodo, leader_legajo=body.aprobador_legajo)
    return {"ok": True, "estado": "RECHAZADO"}


# ==========================================================
# Reabrir parte (líder — vuelve a BORRADOR)
# ==========================================================

@router.post("/parte/{legajo}/{periodo}/reabrir")
async def reabrir_parte(legajo: str, periodo: str, body: AprobarBody):
    parte = db.get_parte(legajo, periodo)
    if not parte:
        raise HTTPException(404, "Parte no encontrado.")
    if parte["estado"] != "APROBADO":
        raise HTTPException(400, f"Solo se pueden reabrir partes APROBADOS. Estado: {parte['estado']}.")
    person = db.get_person(legajo)
    aprobador = str(body.aprobador_legajo).strip()
    if not person or (aprobador not in db.SUPER_LIDERES and str(person.get("leader_legajo","")).strip() != aprobador):
        raise HTTPException(403, "No tenés permiso para reabrir este parte.")
    db.update_parte_estado(legajo, periodo, "BORRADOR",
                           rejection_comment=None,
                           clear_approved=True)
    # Invalida caché del líder directo del empleado (no del super líder que reabrió)
    _invalidar_parte(legajo, periodo)
    return {"ok": True, "estado": "BORRADOR"}


# ==========================================================
# Bitácora del empleado
# ==========================================================

@router.get("/bitacora/{legajo}")
async def get_bitacora(legajo: str):
    key    = _k_bitacora(legajo)
    cached = cache.get(key)
    if cached is not None:
        return cached
    partes = db.list_bitacora(legajo)
    result_partes = []
    for p in partes:
        start, end = db.period_bounds(p["periodo"])
        result_partes.append({
            **p,
            "periodo_display": db.period_display(p["periodo"]),
            "periodo_inicio":  start.isoformat(),
            "periodo_fin":     end.isoformat(),
        })
    result = {"legajo": legajo, "partes": result_partes}
    cache.set(key, result, ttl=_TTL_BITACORA)
    return result


# ==========================================================
# Equipo del líder
# ==========================================================

@router.get("/equipo/{leader_legajo}/pendientes")
async def get_pendientes(leader_legajo: str):
    key    = _k_pendientes(leader_legajo)
    cached = cache.get(key)
    if cached is not None:
        return cached
    rows = db.list_pendientes_lider(leader_legajo)
    result_rows = []
    for r in rows:
        start, end = db.period_bounds(r["periodo"])
        result_rows.append({
            **r,
            "periodo_display": db.period_display(r["periodo"]),
            "periodo_inicio":  start.isoformat(),
            "periodo_fin":     end.isoformat(),
        })
    result = {"leader_legajo": leader_legajo, "pendientes": result_rows}
    cache.set(key, result, ttl=_TTL_PENDIENTE)
    return result


@router.get("/equipo/{leader_legajo}/partes")
async def get_team_partes(leader_legajo: str, periodo: Optional[str] = None):
    key    = f"rrhh:team_partes:{leader_legajo}:{periodo or 'all'}"
    cached = cache.get(key)
    if cached is not None:
        return cached
    rows = db.list_team_partes(leader_legajo, periodo)
    result_rows = []
    for r in rows:
        start, end = db.period_bounds(r["periodo"])
        result_rows.append({
            **r,
            "periodo_display": db.period_display(r["periodo"]),
            "periodo_inicio":  start.isoformat(),
            "periodo_fin":     end.isoformat(),
        })
    result = {"leader_legajo": leader_legajo, "partes": result_rows}
    cache.set(key, result, ttl=_TTL_PENDIENTE)
    return result


# ==========================================================
# Consolidado
# ==========================================================

@router.get("/consolidado/{leader_legajo}/{periodo}")
async def get_consolidado(leader_legajo: str, periodo: str):
    key    = _k_consolidado(leader_legajo, periodo)
    cached = cache.get(key)
    if cached is not None:
        return cached
    data   = db.get_consolidado(leader_legajo, periodo)
    start, end = db.period_bounds(periodo)
    result = {
        "leader_legajo":   leader_legajo,
        "periodo":         periodo,
        "periodo_display": db.period_display(periodo),
        "periodo_inicio":  start.isoformat(),
        "periodo_fin":     end.isoformat(),
        "empleados":       data,
    }
    cache.set(key, result, ttl=_TTL_CONSOL)
    return result


# ==========================================================
# Consolidado — Excel (no se cachea, se genera on-demand)
# ==========================================================

@router.get("/consolidado/{leader_legajo}/{periodo}/excel")
async def download_consolidado_excel(leader_legajo: str, periodo: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Usar consolidado cacheado si existe
    key    = _k_consolidado(leader_legajo, periodo)
    cached = cache.get(key)
    if cached:
        data   = cached["empleados"]
        leader = db.get_person(leader_legajo)
        start, end = db.period_bounds(periodo)
        disp   = cached["periodo_display"]
    else:
        data   = db.get_consolidado(leader_legajo, periodo)
        leader = db.get_person(leader_legajo)
        start, end = db.period_bounds(periodo)
        disp   = db.period_display(periodo)

    wb = Workbook()

    # ── Estilos ───────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    SUBHDR_FILL = PatternFill("solid", fgColor="2D5BA6")
    ALT_FILL    = PatternFill("solid", fgColor="F0F4FF")
    TOTAL_FILL  = PatternFill("solid", fgColor="E8F0FE")
    WH          = Font(bold=True, color="FFFFFF")
    BLD         = Font(bold=True)
    CENTER      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Hoja 1: Resumen ───────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Consolidado de Guardias — {disp}"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = HEADER_FILL
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:J2")
    leader_name = leader["nombre"] if leader else leader_legajo
    ws["A2"] = f"Líder: {leader_name}   |   Período: {start.strftime('%d/%m/%Y')} → {end.strftime('%d/%m/%Y')}"
    ws["A2"].font      = Font(size=11, color="FFFFFF")
    ws["A2"].fill      = SUBHDR_FILL
    ws["A2"].alignment = CENTER

    headers = ["Legajo","Nombre","Función","Estado","G","F","D","HO","HV (hs)","HE (hs)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = WH; c.fill = SUBHDR_FILL; c.alignment = CENTER; c.border = BORDER

    for i, emp in enumerate(data):
        row = i + 4
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col, val in enumerate(
            [emp["legajo"], emp["nombre"], emp["funcion"], emp["estado"],
             emp["G"], emp["F"], emp["D"], emp["HO"], emp["HV"], emp["HE"]], 1
        ):
            c = ws.cell(row=row, column=col, value=val)
            c.border = BORDER
            c.alignment = CENTER if col > 3 else Alignment(vertical="center")
            if i % 2 == 0: c.fill = fill

    total_row = len(data) + 4
    ws.merge_cells(f"A{total_row}:D{total_row}")
    ws.cell(row=total_row, column=1, value="TOTAL").font = BLD
    for col_idx, key_t in enumerate(["G","F","D","HO","HV","HE"], 5):
        c = ws.cell(row=total_row, column=col_idx, value=round(sum(e[key_t] for e in data), 2))
        c.font = BLD; c.fill = TOTAL_FILL; c.alignment = CENTER; c.border = BORDER

    widths = [10,30,20,14,6,6,6,6,10,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 20

    # ── Hoja 2: Detalle (acordeón por empleado) ──────────
    ws2 = wb.create_sheet("Detalle")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = f"Detalle por empleado — {disp}"
    ws2["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill      = HEADER_FILL
    ws2["A1"].alignment = CENTER

    det_row = 2
    for emp in data:
        ws2.merge_cells(f"A{det_row}:H{det_row}")
        ws2[f"A{det_row}"] = (
            f"  {emp['nombre']}  |  Legajo {emp['legajo']}"
            f"  |  {emp['funcion']}  |  Estado: {emp['estado']}"
        )
        ws2[f"A{det_row}"].font = Font(bold=True, color="FFFFFF")
        ws2[f"A{det_row}"].fill = SUBHDR_FILL
        ws2.row_dimensions[det_row].height = 18
        det_row += 1

        for col, h in enumerate(["Fecha","G","F","D","HO","HV","HE","Comentario"], 1):
            c = ws2.cell(row=det_row, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4472C4")
            c.alignment = CENTER; c.border = BORDER
        det_row += 1

        for j, dia in enumerate(emp["dias"]):
            f_disp = f"{dia['fecha'][8:]}/{dia['fecha'][5:7]}/{dia['fecha'][:4]}"
            alt = PatternFill("solid", fgColor="EEF2FF") if j % 2 == 0 else PatternFill()
            for col, val in enumerate([
                f_disp,
                "✓" if "G"  in dia["tipos"] else "",
                "✓" if "F"  in dia["tipos"] else "",
                "✓" if "D"  in dia["tipos"] else "",
                "✓" if "HO" in dia["tipos"] else "",
                dia["HV"] or "", dia["HE"] or "",
                dia.get("comentario",""),
            ], 1):
                c = ws2.cell(row=det_row, column=col, value=val)
                c.alignment = CENTER if col < 8 else Alignment(vertical="center")
                c.border = BORDER; c.fill = alt
            det_row += 1

        for col, val in enumerate(
            ["Totales", emp["G"], emp["F"], emp["D"], emp["HO"],
             round(emp["HV"],2), round(emp["HE"],2), ""], 1
        ):
            c = ws2.cell(row=det_row, column=col, value=val)
            c.font = BLD; c.fill = TOTAL_FILL; c.alignment = CENTER; c.border = BORDER
        det_row += 2

    ws2.column_dimensions["A"].width = 14
    for col_l in ["B","C","D","E","F","G"]:
        ws2.column_dimensions[col_l].width = 7
    ws2.column_dimensions["H"].width = 30
    ws2.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"consolidado_{leader_legajo}_{periodo}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
