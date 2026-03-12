#!/usr/bin/env python3
# ==========================================================
# ingest/parser_parte_diario.py
#
# Parser del "Parte Diario de Operaciones de Torre" (.xlsm)
#
# Transforma las hojas "Día 1", "Día 2", ... "Día N" en filas
# con el mismo esquema del histórico.
#
# Columnas que produce (subset acordado del histórico):
#
#   EVENTO (repetido en cada fila de actividad):
#     event_id, well_legal_name, well_id, event_type, event_code,
#     event_objective_1, event_objective_2, status_end,
#     date_ops_start, date_ops_end, contractor_name, rig_name
#
#   ACTIVIDAD (una fila por actividad):
#     step_no, date_report, time_from, time_to, activity_duration,
#     activity_class, activity_code, activity_phase, activity_subcode,
#     expr1, entity_type, date_time_off_location, date_rig_pickup
#
# Lógica status_end:
#   - Último paso SIN time_to → "EN_CURSO"
#   - Último paso CON time_to → "COMPLETADO"
#   - Puede sobreescribirse con force_status
# ==========================================================

from __future__ import annotations

import hashlib
import re
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd


# ==========================================================
# Mapeos
# ==========================================================

_TIPO_OP_MAP: dict[str, tuple[str, str]] = {
    "pulling":           ("INTERVENCION SSPP", "INT"),
    "intervencion":      ("INTERVENCION SSPP", "INT"),
    "intervencion sspp": ("INTERVENCION SSPP", "INT"),
    "workover":          ("REPARACION WO",     "REP"),
    "reparacion wo":     ("REPARACION WO",     "REP"),
    "reparacion":        ("REPARACION WO",     "REP"),
    "terminacion":       ("TERMINACION",       "TER"),
}

_ACTIVITY_CLASS_MAP: dict[str, str] = {
    "OPERA": "P",   # Productivo / Operativo
    "DTM":   "U",   # Traslado
    "CLIMA": "L",   # Clima
    "MANT":  "N",   # Mantenimiento
    "ESP":   "E",   # Espera
    "IND":   "A",   # Indisponible
}

# Fila / columna de los campos en cada hoja "Día N"
_ROW_FECHA    = 6;  _COL_FECHA    = 5   # E6
_ROW_RIG      = 7;  _COL_RIG      = 5   # E7
_ROW_EVENT_ID = 7;  _COL_EVENT_ID = 11  # K7
_ROW_TIPO_OP  = 8;  _COL_TIPO_OP  = 5   # E8
_ROW_POZO     = 8;  _COL_POZO     = 11  # K8
_ROW_CLIENTE  = 10; _COL_CLIENTE  = 5   # E10

_DATA_START = 30   # primera fila de actividades
_DATA_MAX   = 130  # límite superior del scan (suficiente para cualquier día)

# Columnas del bloque de actividades (1-based)
_C_TIME_FROM   = 2   # B → INICIO
_C_TIME_TO     = 3   # C → FIN
_C_DURATION    = 4   # D → DURACIÓN
_C_ACT_CODE    = 5   # E → CÓDIGO OPERATIVO
_C_ACT_SUBCODE = 6   # F → SUB COD.
_C_TARIFA      = 7   # G → TARIFA
_C_DESC        = 13  # M → DESCRIPCIÓN


# ==========================================================
# Helpers
# ==========================================================

def _well_id(name: str) -> str:
    return hashlib.sha1(name.strip().upper().encode()).hexdigest()[:10]


def _map_event_type(tipo: str) -> tuple[str, str]:
    key = (tipo or "").strip().lower()
    for k, v in _TIPO_OP_MAP.items():
        if k in key:
            return v
    return (tipo.upper(), tipo[:3].upper())


def _cell_str(ws, row: int, col: int) -> str | None:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else None


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_time(raw) -> str | None:
    """
    Normaliza el tiempo leído de la celda a 'HH:MM'.
    Acepta:
      - str '14:00:00' o '14:00'
      - timedelta (para medianoche y otras)
    """
    if raw is None:
        return None
    if isinstance(raw, timedelta):
        total = int(raw.total_seconds()) % 86400
        return f"{total // 3600:02d}:{(total % 3600) // 60:02d}"
    s = str(raw).strip()
    if s.lower() in ("ingrese hs", "ingrese"):
        return None
    # 'HH:MM:SS' o 'HH:MM'
    m = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", s)
    if m:
        return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    # '1 day, H:MM:SS' (medianoche en openpyxl)
    m2 = re.match(r"^(\d+) day", s)
    if m2:
        # medianoche → 00:00
        return "00:00"
    return None


def _to_iso(d: date | None, hhmm: str | None) -> str | None:
    if not d or not hhmm:
        return None
    try:
        h, m = map(int, hhmm.split(":"))
        return datetime(d.year, d.month, d.day, h, m).isoformat()
    except Exception:
        return None


# ==========================================================
# Leer Información Inicial (hoja dedicada + fallback Día 1)
# ==========================================================

def _leer_info_inicial(wb: openpyxl.Workbook) -> dict:
    """
    Lee los campos del evento desde hoja 'Información Inicial'.
    Para event_id y datos faltantes, usa la primera hoja 'Día N'.

    Offsets confirmados en Información Inicial (columnas 1-based):
      Fila  7: col4=Equipo          col6=Tipo de Operación
      Fila  8: col4=Contrato        col6=Cliente
      Fila 10: col4=Pozo            col6=Cód. de Intervención (puede ser None)
      Fila 13: col5='Objetivo'      col6=valor
      Fila 14: col2=Tipo de Intervención  col4=valor
    """
    info = {
        "rig_name": None, "contractor_name": None,
        "well_legal_name": None, "event_id": None,
        "tipo_operacion": None,
        "event_objective_1": None, "event_objective_2": None,
    }

    if "Información Inicial" in wb.sheetnames:
        ws = wb["Información Inicial"]
        info["rig_name"]          = _cell_str(ws, 7,  4)
        info["tipo_operacion"]    = _cell_str(ws, 7,  6)
        info["contractor_name"]   = _cell_str(ws, 8,  6)
        info["well_legal_name"]   = _cell_str(ws, 10, 4)
        info["event_id"]          = _cell_str(ws, 10, 6)
        info["event_objective_1"] = _cell_str(ws, 14, 4)
        info["event_objective_2"] = _cell_str(ws, 13, 6)

    # Fallback desde primera hoja Día N para campos faltantes
    dia_sheets = sorted(
        [s for s in wb.sheetnames if re.match(r"^[Dd]ía\s+\d+$", s)],
        key=lambda s: int(re.search(r"\d+", s).group()),
    )
    if dia_sheets:
        ws1 = wb[dia_sheets[0]]
        if not info["event_id"]:
            info["event_id"] = _cell_str(ws1, _ROW_EVENT_ID, _COL_EVENT_ID)
        if not info["well_legal_name"]:
            info["well_legal_name"] = _cell_str(ws1, _ROW_POZO, _COL_POZO)
        if not info["rig_name"]:
            info["rig_name"] = _cell_str(ws1, _ROW_RIG, _COL_RIG)
        if not info["tipo_operacion"]:
            info["tipo_operacion"] = _cell_str(ws1, _ROW_TIPO_OP, _COL_TIPO_OP)
        if not info["contractor_name"]:
            info["contractor_name"] = _cell_str(ws1, _ROW_CLIENTE, _COL_CLIENTE)

    return info


# ==========================================================
# Leer actividades de un día
# ==========================================================

def _leer_actividades_dia(ws, fecha: date) -> list[dict]:
    """
    Lee las filas de actividades de una hoja 'Día N'.
    Para a los 3 filas vacías consecutivas o al detectar 'OBSERVACIONES'.
    """
    actividades: list[dict] = []
    consecutive_empty = 0

    for r in range(_DATA_START, _DATA_MAX):
        tf_raw  = ws.cell(r, _C_TIME_FROM).value
        tt_raw  = ws.cell(r, _C_TIME_TO).value
        dur_raw = ws.cell(r, _C_DURATION).value
        cod_raw = ws.cell(r, _C_ACT_CODE).value
        sub_raw = ws.cell(r, _C_ACT_SUBCODE).value
        tar_raw = ws.cell(r, _C_TARIFA).value
        dsc_raw = ws.cell(r, _C_DESC).value

        # Fin de sección
        if dsc_raw and "observaciones" in str(dsc_raw).lower():
            break

        # Detectar fila real de actividad
        has_time = tf_raw is not None and not (
            isinstance(tf_raw, str) and "ingrese" in tf_raw.lower()
        )
        has_dur  = dur_raw is not None and not (
            isinstance(dur_raw, str) and "ingrese" in dur_raw.lower()
        )

        if not has_time:
            # Contar vacíos para corte anticipado
            if all(ws.cell(r, c).value is None for c in range(1, 15)):
                consecutive_empty += 1
                if actividades and consecutive_empty >= 3:
                    break
            else:
                consecutive_empty = 0
            continue

        consecutive_empty = 0

        tf_str = _normalize_time(tf_raw)
        tt_str = _normalize_time(tt_raw)

        # La última fila puede no tener time_to (intervención en curso)
        incomplete = tt_str is None

        try:
            duration = round(float(dur_raw), 4) if has_dur else None
        except (TypeError, ValueError):
            duration = None

        actividades.append({
            "date_report":       fecha.isoformat(),
            "time_from":         _to_iso(fecha, tf_str),
            "time_to":           _to_iso(fecha, tt_str) if not incomplete else None,
            "activity_duration": duration,
            "activity_code":     str(cod_raw).strip().upper() if cod_raw else None,
            "activity_phase":    str(sub_raw).strip().upper() if sub_raw else None,
            "activity_subcode":  str(tar_raw).strip()         if tar_raw else None,
            "expr1":             str(dsc_raw).strip()         if dsc_raw else None,
            "activity_class":    _ACTIVITY_CLASS_MAP.get(
                                     str(cod_raw).strip().upper() if cod_raw else "", "U"
                                 ),
            "_incomplete":       incomplete,
        })

    return actividades


# ==========================================================
# Parser principal
# ==========================================================

def parsear_parte_diario(
    path: str | Path,
    force_status: str | None = None,
) -> pd.DataFrame:
    """
    Parsea un archivo .xlsm de Parte Diario de Operaciones de Torre.

    Args:
        path:         Ruta al archivo .xlsm / .xlsx
        force_status: Sobreescribe el status_end automático.
                      Valores: "EN_CURSO", "COMPLETADO", "SUSPENDIDO"

    Returns:
        pd.DataFrame con el esquema del histórico. Una fila por actividad.
        Vacío si el archivo no tiene actividades válidas.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ── Datos del evento ──────────────────────────────────
    info = _leer_info_inicial(wb)
    well_name  = info["well_legal_name"] or ""
    event_id   = info["event_id"]        or ""
    rig_name   = info["rig_name"]        or ""
    contractor = info["contractor_name"] or ""
    tipo_op    = info["tipo_operacion"]  or ""
    obj1       = info["event_objective_1"] or ""
    obj2       = info["event_objective_2"] or ""

    event_type, event_code = _map_event_type(tipo_op)
    w_id = _well_id(well_name) if well_name else ""

    # ── Hojas Día N ───────────────────────────────────────
    dia_sheets = sorted(
        [s for s in wb.sheetnames if re.match(r"^[Dd]ía\s+\d+$", s)],
        key=lambda s: int(re.search(r"\d+", s).group()),
    )

    all_acts: list[dict] = []
    date_start: date | None = None
    date_end:   date | None = None
    last_incomplete = False

    for sheet_name in dia_sheets:
        ws = wb[sheet_name]
        fecha = _parse_date(ws.cell(_ROW_FECHA, _COL_FECHA).value)
        if fecha is None:
            continue

        acts = _leer_actividades_dia(ws, fecha)
        if not acts:
            continue

        if date_start is None:
            date_start = fecha
        date_end = fecha
        last_incomplete = acts[-1].get("_incomplete", False)
        all_acts.extend(acts)

    if not all_acts:
        return pd.DataFrame()

    # ── Status_end ────────────────────────────────────────
    if force_status:
        status_end = force_status.upper()
    else:
        status_end = "EN_CURSO" if last_incomplete else "COMPLETADO"

    # ── Construir DataFrame ───────────────────────────────
    rows = []
    for step_no, act in enumerate(all_acts, start=1):
        rows.append({
            # Evento
            "event_id":               event_id,
            "well_legal_name":        well_name,
            "well_id":                w_id,
            "event_type":             event_type,
            "event_code":             event_code,
            "event_objective_1":      obj1,
            "event_objective_2":      obj2,
            "status_end":             status_end,
            "date_ops_start":         date_start.isoformat() if date_start else None,
            "date_ops_end":           date_end.isoformat()   if date_end   else None,
            "contractor_name":        contractor,
            "rig_name":               rig_name,
            # Actividad
            "step_no":                step_no,
            "date_report":            act["date_report"],
            "time_from":              act["time_from"],
            "time_to":                act["time_to"],
            "activity_duration":      act["activity_duration"],
            "activity_class":         act["activity_class"],
            "activity_code":          act["activity_code"],
            "activity_phase":         act["activity_phase"],
            "activity_subcode":       act["activity_subcode"],
            "expr1":                  act["expr1"],
            "entity_type":            "DAILY",
            "date_time_off_location": None,
            "date_rig_pickup":        None,
        })

    return pd.DataFrame(rows)


# ==========================================================
# Merge con histórico
# ==========================================================

def merge_con_historico(
    df_historico: pd.DataFrame,
    df_nuevo: pd.DataFrame,
) -> pd.DataFrame:
    """
    Combina el histórico con las filas del parte nuevo.

    - Elimina del histórico las filas del mismo event_id (evita duplicados).
    - Concatena las filas nuevas.
    - Ordena por date_ops_start, event_id, step_no.

    Si el evento estaba EN_CURSO en el histórico, queda reemplazado
    con el estado actualizado del nuevo parte.
    """
    if df_nuevo.empty:
        return df_historico

    event_ids = df_nuevo["event_id"].unique().tolist()

    if not df_historico.empty and "event_id" in df_historico.columns:
        df_historico = df_historico[
            ~df_historico["event_id"].isin(event_ids)
        ].copy()

    # Alinear columnas
    for col in df_nuevo.columns:
        if col not in df_historico.columns:
            df_historico[col] = None
    for col in df_historico.columns:
        if col not in df_nuevo.columns:
            df_nuevo = df_nuevo.copy()
            df_nuevo[col] = None

    df_merged = pd.concat(
        [df_historico, df_nuevo[df_historico.columns]],
        ignore_index=True,
    )

    sort_cols = [c for c in ["date_ops_start", "event_id", "step_no"]
                 if c in df_merged.columns]
    if sort_cols:
        df_merged = df_merged.sort_values(
            sort_cols, na_position="last"
        ).reset_index(drop=True)

    return df_merged


# ==========================================================
# Standalone / prueba rápida
# ==========================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Uso: python parser_parte_diario.py <ruta_al_xlsm>")
        sys.exit(1)

    df = parsear_parte_diario(sys.argv[1])

    if df.empty:
        print("⚠️  Sin actividades parseadas.")
        sys.exit(0)

    r0 = df.iloc[0]
    print(f"\n✅  {len(df)} actividades parseadas\n")
    print("── EVENTO ────────────────────────────────────────────")
    for col in ["event_id", "well_legal_name", "well_id", "event_type",
                "event_code", "event_objective_1", "status_end",
                "date_ops_start", "date_ops_end", "contractor_name", "rig_name"]:
        print(f"  {col:<25} {r0[col]}")

    print("\n── ACTIVIDADES (muestra) ──────────────────────────────")
    print(f"  {'paso':>4}  {'fecha':12}  {'inicio':6}  {'fin':6}  {'dur':5}  {'cod':6}  descripción")
    for _, r in df.head(5).iterrows():
        tf = str(r.time_from)[11:16] if r.time_from else "?"
        tt = str(r.time_to)[11:16]   if r.time_to   else "?"
        print(f"  {int(r.step_no):>4}  {r.date_report}  {tf}  {tt}  "
              f"{(r.activity_duration or 0):>5.2f}  {(r.activity_code or ''):6}  "
              f"{str(r.expr1)[:50]}")
    print("  ...")
    for _, r in df.tail(2).iterrows():
        tf = str(r.time_from)[11:16] if r.time_from else "?"
        tt = str(r.time_to)[11:16]   if r.time_to   else "?"
        print(f"  {int(r.step_no):>4}  {r.date_report}  {tf}  {tt}  "
              f"{(r.activity_duration or 0):>5.2f}  {(r.activity_code or ''):6}  "
              f"{str(r.expr1)[:50]}")

    print("\n── RESUMEN POR DÍA ────────────────────────────────────")
    for d, n in df.groupby("date_report").size().items():
        print(f"  {d}  →  {n} actividades")

    print(f"\n  status_end final: {df['status_end'].iloc[-1]}")
